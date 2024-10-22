import comtypes.client
import math
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl import load_workbook
import pandas as pd

ProgramPath = r"C:\Program Files\Computers and Structures\ETABS 22\ETABS.exe"
helper = comtypes.client.CreateObject('ETABSv1.Helper')
helper = helper.QueryInterface(comtypes.gen.ETABSv1.cHelper)


#create API helper object
ETABSObject = comtypes.client.GetActiveObject("CSI.ETABS.API.ETABSObject")
SapModel = ETABSObject.SapModel

########################################################################################################################
##############--------------------------------Retrieve Piers Forces First--------------------------------###############
########################################################################################################################

#retrieve all of the pier labels in the model

NumberTables = 1
TableKey = []
TableName = []
ImportType = []

# this shows all the available tables that can be accessed
x = SapModel.DatabaseTables.GetAvailableTables(NumberTables,TableKey, TableName, ImportType)
# LoadCaseList = ['ELF X', 'ELF Y']

LoadCaseList = ['EQXTN', 'EQXTP', 'EQYTN', 'EQYTP']
UA_DFs = [] #Create an empty list to store the UA DFs
Demo_DFs = [] #Create an empty list to store the Demo DFs
# x = SapModel.DatabaseTables.SetLoadCasesSelectedForDisplay(LoadCaseList)
x = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()





############################

# TableKey = 'Element Forces - Braces'
# FieldKeyList = []

# # set the group you want the results for, you can pick either 'All', 'Left Nodes', 'Right Nodes'
# GroupName = 'All'
#
# TableVersion = 1
# FieldsKeysIncluded = []
# NumberRecords = 1
# TableData = []
#
# BraceForces = SapModel.DatabaseTables.GetTableforDisplayArray(TableKey, FieldKeyList, GroupName, TableVersion, FieldsKeysIncluded, NumberRecords, TableData)

def cut_list(lst, chunk_size):
    return [lst[i:i + chunk_size] for i in range(0, len(lst), chunk_size)]

# BraceLabels = BraceForces[2]

# print(len(BraceLabels))
# IndividualBraceForces = cut_list(BraceForces[4], len(BraceLabels))
# selectedBraceForces =IndividualBraceForces[0::6]
#
# # Creating the DataFrame
# df = pd.DataFrame(selectedBraceForces, columns=BraceLabels)
# df[['P', 'V2', 'V3', 'T', 'M2', 'M3']] = df[['P', 'V2', 'V3', 'T', 'M2', 'M3']].astype(float)
#
# file_path = 'BraceForcesLACC.xlsx'
#
# SheetName = f'UA Brace Forces'
# SheetName = f'Demo Braces Forces'

# UA_DFs = []
# Demo_DFs = []
#
# if "UA" in SheetName:
#     UA_DFs.append(df)
#     print('Load DFs in UA')
# else:
#     Demo_DFs.append(df)
#     print('Load DFs in Demo')
#
# # Try to load the workbook and check if the sheet exists
# try:
#     workbook = openpyxl.load_workbook(file_path)
#     if SheetName in workbook.sheetnames:
#         print(f"Sheet '{SheetName}' already exists. No changes made.")
#     else:
#         # Write to the Excel file if the sheet does not exist
#         with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
#             df.to_excel(writer, sheet_name=SheetName, index=False)
#         print(f"Data written to {SheetName} sheet.")
# except FileNotFoundError:
#     # If the file does not exist, create it and write the data
#     with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
#         df.to_excel(writer, sheet_name=SheetName, index=False)
#     print(f"File '{file_path}' created and data written to '{SheetName}' sheet.")

#For Each Load Case, let's extract the values and dataframes we care about.
for load in LoadCaseList:
    x = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
    LoadCase = [load]
    x = SapModel.DatabaseTables.SetLoadCombinationsSelectedForDisplay(LoadCase)

    TableKey = 'Element Forces - Braces'
    FieldKeyList = []

    # set the group you want the results for, you can pick either 'All', 'Left Nodes', 'Right Nodes'
    GroupName = 'All'

    TableVersion = 1
    FieldsKeysIncluded = []
    NumberRecords = 1
    TableData = []

    BraceForcesUA = SapModel.DatabaseTables.GetTableforDisplayArray(TableKey, FieldKeyList, GroupName, TableVersion, FieldsKeysIncluded, NumberRecords, TableData)

    BraceLabels = BraceForcesUA[2]
    IndividualBraceForces = cut_list(BraceForcesUA[4], len(BraceLabels))
    # selectedBraceForces = IndividualBraceForces[0::6]


    # Creating the DataFrame
    df = pd.DataFrame(IndividualBraceForces, columns=BraceLabels)
    df[['Station','P', 'V2', 'V3', 'T', 'M2', 'M3']] = df[['Station','P', 'V2', 'V3', 'T', 'M2', 'M3']].astype(float)
    filtered_df = df[(df['Station'] == 0) & (df['StepType'] == 'Max')]



    #Writing Pier Forces to Excel

    # Load the Excel file
    # Delete the UA Pier or Demo Pier Sheet to write new values in. These values will be read to compare pier forces.

    file_path = 'BraceForcesLACC.xlsx'

    SheetName = f'UA Brace Forces {load}'
    # SheetName = f'Temp Brace Forces {load}'

    if "UA" in SheetName:
        UA_DFs.append(filtered_df)
        print('Load DFs in UA')
    else:
        Demo_DFs.append(filtered_df)

    # Try to load the workbook and check if the sheet exists
    try:
        workbook = openpyxl.load_workbook(file_path)
        if SheetName in workbook.sheetnames:
            print(f"Sheet '{SheetName}' already exists. No changes made.")
        else:
            # Write to the Excel file if the sheet does not exist
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
                filtered_df.to_excel(writer, sheet_name=SheetName, index=False)
            print(f"Data written to {SheetName} sheet.")
    except FileNotFoundError:
        # If the file does not exist, create it and write the data
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            filtered_df.to_excel(writer, sheet_name=SheetName, index=False)
        print(f"File '{file_path}' created and data written to '{SheetName}' sheet.")


########################################################################################################################
                             #LOAD UP THE STORED DATA into correct List of Dataframes.
########################################################################################################################



#LOAD UP THE STORED DATA into correct List of Dataframes.
for load in LoadCaseList:
    if "UA" in SheetName:
        Demo_DFs.append(pd.read_excel(file_path, sheet_name=f'Temp Brace Forces {load}'))
        print(f'Successfully loaded Temp Pier Forces {load}')
    else:
        UA_DFs.append(pd.read_excel(file_path, sheet_name=f'UA Brace Forces {load}'))
        print(f'Successfully loaded UA Pier Forces {load}')

Demo_DFs_MAX = []
UA_DFs_MAX = []

Demo_DFs_MAX = []
UA_DFs_MAX = []

#This picks the abs max of shear force for V2

for Demo_DF in Demo_DFs:
    Demo_DFs_MAX.append(Demo_DF.groupby(['Story', 'UniqueName'], as_index=False).apply(lambda group: group.loc[group['P'].abs().idxmax()]))

for UA_DF in UA_DFs:
    UA_DFs_MAX.append(UA_DF.groupby(['Story', 'UniqueName'], as_index=False).apply(lambda group: group.loc[group['P'].abs().idxmax()]))

# List to store the matching index pairs
matching_indices = []
#
# # Loop through each row in PierForces_UA_Level245_df
for ua_index, ua_row in UA_DFs_MAX[0].iterrows():
    # Get the 'Pier' value from PierForces_UA_Level245_df
    ua_pier_value = ua_row['UniqueName']
    # Find the corresponding row in PierForces_Demo_Level245_df
    demo_index = Demo_DFs_MAX[0][Demo_DFs_MAX[0]['UniqueName'] == ua_pier_value].index
    # If a match is found, store the indices
    if not demo_index.empty:
        matching_indices.append([ua_index, int(demo_index[0])])
        #Store none value if it is empty
    else:
        matching_indices.append([ua_index, None])



comparison_data = []


group_name = 'Failing Braces'
try:
    x = SapModel.GroupDef.Delete(group_name)
    print("deleted group")
except:
    x = SapModel.GroupDef.SetGroup_1(group_name)
    print("Created a group")


x = SapModel.GroupDef.SetGroup_1(group_name)
print("Created a group")


for ua_index, demo_index in matching_indices:
    # Get the 'UniqueName' value from the UA DataFrame

    brace_value = UA_DFs_MAX[0].at[ua_index, 'UniqueName']
    brace_level = UA_DFs_MAX[0].at[ua_index, 'Story']
    x = SapModel.FrameObj.GetSection(brace_value)
    brace_size = x[0]
    print([ua_index, demo_index, brace_value])
    if demo_index is not None:  # Ensure there is a corresponding match
        UA_V2_value = 0
        UA_M3_value = 0
        UA_V2_index = 0
        UA_M3_index = 0
        V2_Comparison = []
        M3_Comparison = []

        Demo_V2_value = 0
        Demo_M3_value = 0
        Demo_V2_index = 0
        Demo_M3_index = 0

        # Review of Axial
        for i in range(len(UA_DFs_MAX)):
            #We are only reviewing the values with the highest result
            if abs(Demo_DFs_MAX[i].at[demo_index, 'P']) > Demo_V2_value:
                Demo_V2_value = abs(Demo_DFs_MAX[i].at[demo_index, 'P'])
                Demo_V2_index = i

        ua_v2 = abs(UA_DFs_MAX[Demo_V2_index].at[ua_index, 'P'])



        p_comparison = (Demo_V2_value - ua_v2) / ua_v2  # Change in Shear

        if p_comparison >= 0.1: #if the brace has a change by more than 10%, add it to "Failing Brace group"
            x = SapModel.FrameObj.SetGroupAssign(brace_value, group_name)
            message = f"Failing Brace Element {brace_value} added"
            # print(message)


        #m3_comparison = (Demo_M3_value - ua_m3) / ua_m3  # Change in Moment
        load = LoadCaseList[Demo_V2_index]  # The load which controls.

        comparison_data.append({
            'Brace': brace_value,
            'P_Comparison': p_comparison,
            'Load': load,
            'PeB' : ua_v2,
            'PeT' : Demo_V2_value,
            'Story': brace_level,
            'Brace Size': brace_size
        })
    else:
        # If there's no corresponding match, store None or any placeholder, AKA, wall was demo'd
        UA_V2_value_demod = 0
        x = SapModel.FrameObj.GetSection(brace_value)
        brace_size = x[0]

        for i in range(len(UA_DFs_MAX)):
            ua_v2_demod = abs(UA_DFs_MAX[i].at[ua_index, 'P'])
            if abs(ua_v2_demod) > UA_V2_value_demod:
                UA_V2_value_demod = abs(ua_v2_demod)
            else:
                pass

        comparison_data.append({
            'Brace': brace_value,
            'P_Comparison': None,
            'Load': load,
            'PeB': ua_v2,
            'PeT': None,
            'Story': 'Demod',
            'Brace Size': brace_size
        })

# # Create a DataFrame from the comparison data
comparison_df = pd.DataFrame(comparison_data)

# Define the path to the Excel file where you want to save the data
output_file_path = 'BraceForcesLACC.xlsx'
#
# Load the existing workbook
book = load_workbook(output_file_path)

# Write the comparison DataFrame to the specified Excel sheet
with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a') as writer:
    if 'Change' in writer.book.sheetnames:
        del writer.book['Change']
    comparison_df.to_excel(writer, sheet_name='Change', index=False)

print(f"Comparison data has been written to the '{output_file_path}' file in the 'comparison' sheet.")











