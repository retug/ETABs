from beamClass import ETABsBeam
import comtypes.client
import math
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl import load_workbook
import pandas as pd

#Using Comtypes 1.1.14
#boiler plate code to attach to the SAP model
#V2 grabs the wall thickness associated with a pier force.


ProgramPath = r"C:\Program Files\Computers and Structures\ETABS 21\ETABS.exe"
helper = comtypes.client.CreateObject('ETABSv1.Helper')
helper = helper.QueryInterface(comtypes.gen.ETABSv1.cHelper)


#create API helper object
ETABSObject = comtypes.client.GetActiveObject("CSI.ETABS.API.ETABSObject")
SapModel = ETABSObject.SapModel

########################################################################################################################
##############--------------------------------Retrieve Piers Forces First--------------------------------###############
########################################################################################################################
# The model must have been run and results are saved.

#retrieve all of the pier labels in the model

NumberTables = 1
TableKey = []
TableName = []
ImportType = []

# this shows all the available tables that can be accessed
x = SapModel.DatabaseTables.GetAvailableTables(NumberTables,TableKey, TableName, ImportType)
############################

TableKey = 'Pier Label Definitions'
FieldKeyList = []

# set the group you want the results for, you can pick either 'All', 'Left Nodes', 'Right Nodes'
GroupName = 'All'

TableVersion = 1
FieldsKeysIncluded = []
NumberRecords = 1
TableData = []

PierLabels = SapModel.DatabaseTables.GetTableforDisplayArray(TableKey, FieldKeyList, GroupName, TableVersion, FieldsKeysIncluded, NumberRecords, TableData)

#######################################
TableKey = 'Pier Forces'
FieldKeyList = []

# set the group you want the results for, you can pick either 'All', 'Left Nodes', 'Right Nodes'

#Set Load Case for results you want to extract.
LoadCaseList = ['UI - ELF X', 'UI - ELF Y']

UA_DFs = [] #Create an empty list to store the UA DFs
Demo_DFs = [] #Create an empty list to store the Demo DFs


# Function to break apart the Database Table into individual rows of data.
def cut_list(lst, chunk_size):
    return [lst[i:i + chunk_size] for i in range(0, len(lst), chunk_size)]

#For Each Load Case, let's extract the values and dataframes we care about.
for load in LoadCaseList:
    x = SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
    LoadCase = [load]
    x = SapModel.DatabaseTables.SetLoadCasesSelectedForDisplay(LoadCase)
    GroupName = 'All'

    TableVersion = 1
    FieldsKeysIncluded = []
    NumberRecords = 1
    TableData = []

    PierForcesUA = SapModel.DatabaseTables.GetTableforDisplayArray(TableKey, FieldKeyList, GroupName, TableVersion, FieldsKeysIncluded, NumberRecords, TableData)

    PierLabels = PierForcesUA[2]
    IndividualPierForcesUA = cut_list(PierForcesUA[4], len(PierLabels))

    # Creating the DataFrame
    df = pd.DataFrame(IndividualPierForcesUA, columns=PierLabels)
    df[['P', 'V2', 'V3', 'T', 'M2', 'M3']] = df[['P', 'V2', 'V3', 'T', 'M2', 'M3']].astype(float)

    # Filter out to just the bottom most level, 245.58'. Compare shears at the Bottom
    PierForces_UA_Level245_df = df[(df['Story'] == "L01.8 (245.58')") & (df['Location'] == 'Bottom')]

    #Writing Pier Forces to Excel

    # Load the Excel file
    # Delete the UA Pier or Demo Pier Sheet to write new values in. These values will be read to compare pier forces.

    file_path = 'PierForcesLACC.xlsx'

    SheetName = f'UA Pier Forces {load}'
    #SheetName = f'Demo Pier Forces {load}'

    if "UA" in SheetName:
        UA_DFs.append(PierForces_UA_Level245_df)
        print('Load DFs in UA')
    else:
        Demo_DFs.append(PierForces_UA_Level245_df)

    # Try to load the workbook and check if the sheet exists
    try:
        workbook = openpyxl.load_workbook(file_path)
        if SheetName in workbook.sheetnames:
            print(f"Sheet '{SheetName}' already exists. No changes made.")
        else:
            # Write to the Excel file if the sheet does not exist
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
                PierForces_UA_Level245_df.to_excel(writer, sheet_name=SheetName, index=False)
            print(f"Data written to {SheetName} sheet.")
    except FileNotFoundError:
        # If the file does not exist, create it and write the data
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            PierForces_UA_Level245_df.to_excel(writer, sheet_name=SheetName, index=False)
        print(f"File '{file_path}' created and data written to '{SheetName}' sheet.")

##########################################################################################################
#Now we have the data loaded into the following locations UA Pier Forces sheet and Demo Pier Forces sheet.
#Let's compare the two to see values.
############################################################################################################

#LOAD UP THE STORED DATA into correct List of Dataframes.
for load in LoadCaseList:
    if "UA" in SheetName:
        Demo_DFs.append(pd.read_excel(file_path, sheet_name=f'Demo Pier Forces {load}'))
        print(f'Successfully loaded Demo Pier Forces {load}')
    else:
        UA_DFs.append(pd.read_excel(file_path, sheet_name=f'UA Pier Forces {load}'))
        print(f'Successfully loaded UA Pier Forces {load}')

# List to store the matching index pairs
matching_indices = []
#
# # Loop through each row in PierForces_UA_Level245_df
for ua_index, ua_row in UA_DFs[0].iterrows():
    # Get the 'Pier' value from PierForces_UA_Level245_df
    ua_pier_value = ua_row['Pier']
    # Find the corresponding row in PierForces_Demo_Level245_df
    demo_index = Demo_DFs[0][Demo_DFs[0]['Pier'] == ua_pier_value].index
    # If a match is found, store the indices
    if not demo_index.empty:
        matching_indices.append([ua_index, demo_index[0]])
    #Store none value if it is empty
    else:
        matching_indices.append([ua_index, None])


#Let's now write the data of interest back to a dataframe and excel file for final review.

#List to store the comparison data
comparison_data = []

# Loop through the matching indices

#This algorithm works by looking at both shear and moment in the demo condition and picking the max of Shear and Moment
#under the given load combinations, in this example, both X and Y load cases.
#With the max shear and moment in the demo conditions, it then goes and finds the previous shear and moment under the
#corresponding load case in the unaltered condition. The program then calucalted the % increase in shear and moment.
#The largest value is reported in the plot at the coresponding load case.

for ua_index, demo_index in matching_indices:
    # Get the 'Pier Label' value from the UA DataFrame
    pier_value = UA_DFs[0].at[ua_index, 'Pier']

    if demo_index is not None:  # Ensure there is a corresponding match
        UA_V2_value = 0
        UA_M3_value = 0
        UA_V2_index = 0
        UA_M3_index = 0
        V2_Comparison = []
        M3_Comparison = []

        Demo_V2_value = 0
        Demo_M3_value = 0
        Demo_V2_index = 0
        Demo_M3_index = 0

        # Review of Shears
        for i in range(len(UA_DFs)):
            #We are only reviewing the values with the highest result
            if abs(Demo_DFs[i].at[demo_index, 'V2']) > Demo_V2_value:
                Demo_V2_value = Demo_DFs[i].at[demo_index, 'V2']
                Demo_V2_index = i

        #Review of Moments
        for i in range(len(UA_DFs)):
            if abs(Demo_DFs[i].at[demo_index, 'M3']) > Demo_M3_value:
                # UA_M3_value = UA_DFs[i].at[ua_index, 'M3']
                # UA_M3_index = i
                Demo_M3_value = Demo_DFs[i].at[demo_index, 'M3']
                Demo_M3_index = i
        ua_v2 = UA_DFs[Demo_V2_index].at[ua_index, 'V2']
        ua_m3 = UA_DFs[Demo_M3_index].at[ua_index, 'M3']
        v2_comparison = (Demo_V2_value - ua_v2) / ua_v2  # Change in Shear
        #m3_comparison = (Demo_M3_value - ua_m3) / ua_m3  # Change in Moment
        m3_comparison = 0
        if v2_comparison > m3_comparison:
            load = LoadCaseList[Demo_V2_index]  # The load which controls.
            VorM = "Shear"
        else:
            load = LoadCaseList[Demo_M3_index]  # The load which controls.
            VorM = "Moment"

        comparison_data.append({
            'Pier': pier_value,
            'V2_Comparison': v2_comparison,
            'M3_Comparison': m3_comparison,
            'Load': load,
            'VorM' : VorM,
            'VeB' : ua_v2,
            'VeT' : Demo_V2_value
        })
    else:
        # If there's no corresponding match, store None or any placeholder, AKA, wall was demo'd
        UA_V2_value_demod = 0
        print(ua_index)
        for i in range(len(UA_DFs)):
            ua_v2_demod = UA_DFs[i].at[ua_index, 'V2']
            if abs(ua_v2_demod) > UA_V2_value_demod:
                UA_V2_value_demod = abs(ua_v2_demod)
            else:
                pass

        comparison_data.append({
            'Pier': pier_value,
            'V2_Comparison': None,
            'M3_Comparison': None,
            'Load': None,
            'VorM': "Demo'd",
            'VeB': UA_V2_value_demod,
            'VeT': None
        })


# # Create a DataFrame from the comparison data
comparison_df = pd.DataFrame(comparison_data)

# Define the path to the Excel file where you want to save the data
output_file_path = 'PierForcesLACC.xlsx'
#
# Load the existing workbook
book = load_workbook(output_file_path)

# Write the comparison DataFrame to the specified Excel sheet
with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a') as writer:
    if 'Change' in writer.book.sheetnames:
        del writer.book['Change']
    comparison_df.to_excel(writer, sheet_name='Change', index=False)

print(f"Comparison data has been written to the '{output_file_path}' file in the 'comparison' sheet.")
#


####################################Let's now plot the results.########################################################
# Demo walls are dashed, grey
# Walls with DCR over 10% are highlighed in red. DCR change is printed over the red ones.
# Walls with DCR less than 10% are highlighted in green.

#retreive defined walls piers

TableKey = 'Area Assignments - Pier Labels'
FieldKeyList = []

# set the group you want the results for, you can pick either 'All', 'Left Nodes', 'Right Nodes'
GroupName = 'All'

TableVersion = 1
FieldsKeysIncluded = []
NumberRecords = 1
TableData = []

AreaPierLabels = SapModel.DatabaseTables.GetTableforDisplayArray(TableKey, FieldKeyList, GroupName, TableVersion, FieldsKeysIncluded, NumberRecords, TableData)

areaLabels = AreaPierLabels[2]
areaPierInfo = cut_list(AreaPierLabels[4], len(AreaPierLabels[2]))

# Creating the DataFrame
areaLabels_df = pd.DataFrame(areaPierInfo, columns=areaLabels)
#only retreive lower level labels
lowerAreaLabels_df = areaLabels_df[(areaLabels_df['Story'] == "L01.8 (245.58')")]

WallPierandLengths = []

#For all piers, let's now plot a line for the results
for uniqueName, pierName in  zip(lowerAreaLabels_df['UniqueName'],lowerAreaLabels_df['PierName']):

    NumberAreaPoints = 1
    ObjectNamePnts = []
    areaPoints = SapModel.AreaObj.GetPoints(uniqueName, NumberAreaPoints, ObjectNamePnts)
    PropName = ""
    areaProperty = SapModel.AreaObj.GetProperty(uniqueName, PropName)

    WallPropType = 0
    ShellType = 0

    MatProp = ""
    Thickness = 0.00
    Color = 0
    Notes = ""
    GUID = ""
    wallProperty = SapModel.PropArea.GetWall(areaProperty[0], WallPropType, ShellType, MatProp, Thickness, Color, Notes, GUID)
    wallThickness = wallProperty[3] #This is the wall thickness of an individual wall, many walls make up a pier.
    lineX = []
    lineY = []
    for point in areaPoints[1]:
        X = 0
        Y = 0
        Z = 0
        coordinates = SapModel.PointObj.GetCoordCartesian(point, X,Y,Z)
        if coordinates[2]<0.1:
            lineX.append(coordinates[0]/12)
            lineY.append(coordinates[1]/12)


    #Find the row with the correspoding Pier Name in the comparison_df
    row = comparison_df[comparison_df['Pier'] == pierName]
    # Extract the V2 and M3 values
    v2_value = row['V2_Comparison'].values[0]
    m3_value = row['M3_Comparison'].values[0]

    load_text = row['Load'].values[0]
    shear_or_moment = row['VorM'].values[0]
    pier = row['Pier'].values[0]

    pierLength = abs((lineY[1] - lineY[0]) ** 2 + (lineX[1] - lineX[0]) ** 2) ** 0.5
    WallPierandLengths.append([pierLength, pier, wallThickness])

    if v2_value > 0.1 or m3_value > 0.1:
        plt.plot(lineX, lineY, color='red')
        # Calculate the midpoint for lineX and lineY
        mid_point_x = np.mean(lineX)
        mid_point_y = np.mean(lineY)

        # Calculate the maximum value of V2_Comparison and M3_Comparison
        max_value = max(v2_value, m3_value)

        # Add the max_value as text to the plot at the midpoint
        plt.text(mid_point_x+15, mid_point_y, f'{max_value:.2f} \n {load_text} {shear_or_moment} \n {pier}', color='black', fontsize=6, ha='center')
    elif math.isnan(v2_value) or math.isnan(m3_value):
        plt.plot(lineX, lineY, color='grey', linestyle='--')
    else:
        plt.plot(lineX, lineY, color='green')

# Create a dictionary to hold the sums of lengths for each unique string
length_piers = {}

# Iterate over the data and sum the lengths for each unique string
for length, string, wallT in WallPierandLengths:
    key = (string, wallT)
    if key in length_piers:
        length_piers[key] += length
    else:
        length_piers[key] = length

# Convert WallPierandLengths to DataFrame
wall_pier_df = pd.DataFrame(
    [(key[0], key[1], value) for key, value in length_piers.items()],
    columns=['Pier', 'WallT', 'Length']
)
#merge the dataframes with pier labels
comparison_df_merged = pd.merge(comparison_df, wall_pier_df, on='Pier', how='left')

#calculate the temporary root

fprimec = 4000

#This is the root f prime c before and after.
comparison_df_merged['Root F`c Before'] = comparison_df_merged['VeB']/(comparison_df_merged['Length']*12* comparison_df_merged['WallT']*fprimec**0.5)
comparison_df_merged['Root F`c After'] = comparison_df_merged['VeT']/(comparison_df_merged['Length']*12* comparison_df_merged['WallT']*fprimec**0.5)

# Define the path to the Excel file where you want to save the data
output_file_path = 'PierForcesLACC.xlsx'
#
# Load the existing workbook
book = load_workbook(output_file_path)

# Write the comparison DataFrame to the specified Excel sheet
with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a') as writer:
    if 'ChangeV2' in writer.book.sheetnames:
        del writer.book['ChangeV2']
    comparison_df_merged.to_excel(writer, sheet_name='ChangeV2', index=False)

print(f"Comparison data has been written to the '{output_file_path}' file in the 'ChangeV2' sheet.")



# Adding labels and title (optional)
plt.xlabel('X - Feet')
plt.ylabel('Y - Feet')
plt.title('Shear Wall Plot Map')

# Display the plot
plt.show()





